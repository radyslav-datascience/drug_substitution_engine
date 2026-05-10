# =============================================================================
# top_1k_reliability_sales_volume — RUN SUBSTITUTES SUBSET
# =============================================================================
# Призначення: відфільтрувати substitute_shares_power_bi.xlsx так, щоб у ньому
# залишилися пари тільки для тих DRUGS_ID, що пройшли основний фільтр у
# run_filter.py. Тобто DRUGS_ID, наявні у
# outputs/drug_coefficients_power_bi_sales_volume_filtered.xlsx.
#
# Цей скрипт ВИКОНУЄТЬСЯ ПІСЛЯ run_filter.py.
#
# Запуск:
#     python _optional_calculations/top_1k_reliability_sales_volume/run_substitutes_subset.py
# =============================================================================

import logging
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import config  # noqa: E402


def setup_logging() -> Path:
    config.LOGS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = config.LOGS_DIR / f"run_substitutes_{timestamp}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return log_path


def main() -> int:
    log_path = setup_logging()
    t0 = time.time()
    logging.info("=" * 70)
    logging.info("RUN substitutes_subset — start")
    logging.info("=" * 70)

    try:
        # 1. Validate dependencies
        if not config.OUTPUT_ANALYSIS_XLSX.exists():
            raise FileNotFoundError(
                f"Не знайдено фінал основного фільтра: {config.OUTPUT_ANALYSIS_XLSX}\n"
                f"Спочатку запусти run_filter.py."
            )
        if not config.INPUT_SUBSTITUTE_SHARES.exists():
            raise FileNotFoundError(f"Не знайдено вхідний substitute_shares: {config.INPUT_SUBSTITUTE_SHARES}")

        # 2. Load filtered drug list (output of run_filter.py)
        logging.info("[1/3] Loading filtered drug list...")
        df_filtered = pd.read_excel(config.OUTPUT_ANALYSIS_XLSX)
        keep_ids = set(df_filtered["DRUGS_ID"].astype(int).tolist())
        logging.info(f"  filtered drugs: {len(keep_ids):,}")

        # 3. Load full substitute_shares (csv або xlsx — за розширенням)
        logging.info(f"[2/3] Loading substitute_shares from {config.INPUT_SUBSTITUTE_SHARES.name}...")
        if config.INPUT_SUBSTITUTE_SHARES.suffix.lower() == ".csv":
            df_subs_in = pd.read_csv(config.INPUT_SUBSTITUTE_SHARES, sep=";")
        else:
            df_subs_in = pd.read_excel(config.INPUT_SUBSTITUTE_SHARES)
        n0 = len(df_subs_in)
        logging.info(f"  input pairs:    {n0:,}  ({df_subs_in['DRUGS_ID'].nunique():,} unique source drugs)")

        # 4. Subset by DRUGS_ID
        df_subs_out = df_subs_in[df_subs_in["DRUGS_ID"].isin(keep_ids)].copy()
        n_kept = len(df_subs_out)
        n_unique_kept = df_subs_out["DRUGS_ID"].nunique()
        logging.info(f"  subset pairs:   {n_kept:,}  ({n_unique_kept:,} unique source drugs)")
        logging.info(f"  filtered out:   {n0 - n_kept:,} pairs (для DRUGS_ID, що НЕ пройшли основний фільтр)")

        # Inform if some filtered drugs are missing in substitute_shares
        in_subs = set(df_subs_out["DRUGS_ID"].astype(int).unique())
        missing = keep_ids - in_subs
        if missing:
            logging.info(f"  (інфо) {len(missing):,} відфільтрованих препаратів НЕ мають substitute pairs у substitute_shares")
            logging.info(f"         (це нормально — деякі препарати не мали валідних замінників)")

        # 5. Preserve dtypes from input file
        for col in df_subs_out.columns:
            in_dtype = df_subs_in[col].dtype
            if df_subs_out[col].dtype != in_dtype:
                try:
                    df_subs_out[col] = df_subs_out[col].astype(in_dtype)
                except Exception as e:
                    logging.warning(f"  could not cast {col} to {in_dtype}: {e}")

        # 6. Sheet name для xlsx-виходу — стандартизований (CSV input не має sheet'а).
        # Зберігаємо ту саму назву, що використовується у нашому продакшн-файлі
        # results/final/substitute_shares.xlsx — "Substitute Shares".
        if config.INPUT_SUBSTITUTE_SHARES.suffix.lower() == ".xlsx":
            from openpyxl import load_workbook
            in_sheet = load_workbook(config.INPUT_SUBSTITUTE_SHARES, read_only=True).active.title
        else:
            in_sheet = "Substitute Shares"
        logging.info(f"  output sheet name: '{in_sheet}'")

        # 7. Write output xlsx (same dtypes, same sheet name as input)
        logging.info("[3/3] Writing output...")
        config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
        float_cols = [c for c in df_subs_out.columns if str(df_subs_in[c].dtype).startswith("float")]
        with pd.ExcelWriter(config.OUTPUT_SUBSTITUTE_SHARES_XLSX, engine="openpyxl") as writer:
            df_subs_out.to_excel(writer, index=False, sheet_name=in_sheet)
            ws = writer.sheets[in_sheet]
            col_letters = {c: chr(ord("A") + i) for i, c in enumerate(df_subs_out.columns)}
            for c in float_cols:
                letter = col_letters[c]
                for cell in ws[letter][1:]:
                    cell.number_format = "0.000000"

        size_kb = config.OUTPUT_SUBSTITUTE_SHARES_XLSX.stat().st_size / 1024
        logging.info(f"  xlsx written: {config.OUTPUT_SUBSTITUTE_SHARES_XLSX.name} "
                     f"({df_subs_out.shape[0]:,} rows x {df_subs_out.shape[1]} cols, {size_kb:.1f} KB)")

        elapsed = time.time() - t0
        logging.info("")
        logging.info(f"DONE in {elapsed:.1f}s — log: {log_path}")
        logging.info("=" * 70)
        return 0
    except Exception as e:
        logging.exception(f"FAILED: {type(e).__name__}: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
